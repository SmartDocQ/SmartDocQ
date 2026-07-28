import base64
import io
import csv
import openpyxl
import logging
import threading
from flask import Blueprint, request, jsonify
from rank_bm25 import BM25Okapi

from db.chroma import collection
from services.embedding_service import embed_document
from indexing.pipeline import compute_chunk_hash, build_chunk_metadata, _build_contextual_header
from services.bm25_service import tokenize, _bm25_cache, _bm25_lock
from config import CHUNKING_VERSION, INDEX_PIPELINE_VERSION, EMBED_MODEL
from indexing.chunking import estimate_token_count
from utils.table_extraction import flatten_table_for_embedding

logger = logging.getLogger(__name__)

table_edit_bp = Blueprint("table_edit", __name__)

_bm25_timers = {}
_bm25_timers_lock = threading.Lock()

def debounce_bm25_rebuild(doc_id: str, index_version: str):
    def perform_rebuild():
        logger.info("[BM25 Debounced] Starting background rebuild for doc_id=%s version=%s", doc_id, index_version)
        try:
            with _bm25_lock:
                entry = _bm25_cache.get((doc_id, index_version))
                if entry:
                    tokenized = [tokenize(t) for t in entry["texts"]]
                    entry["bm25"] = BM25Okapi(tokenized)
            logger.info("[BM25 Debounced] Finished background rebuild for doc_id=%s version=%s", doc_id, index_version)
        except Exception as err:
            logger.error("[BM25 Debounced] Failed background rebuild for doc_id=%s version=%s: %s", doc_id, index_version, err)
        finally:
            with _bm25_timers_lock:
                _bm25_timers.pop((doc_id, index_version), None)

    with _bm25_timers_lock:
        existing = _bm25_timers.get((doc_id, index_version))
        if existing:
            existing.cancel()
        
        timer = threading.Timer(3.0, perform_rebuild)
        _bm25_timers[(doc_id, index_version)] = timer
        timer.start()

@table_edit_bp.route("/api/internal/document/sync-table", methods=["POST"])
def sync_table():
    """Sync table cell modifications: update binary file bytes, re-embed changed chunks, update Chroma DB."""
    body = request.get_json(silent=True) or {}
    doc_id = body.get("doc_id")
    filename = body.get("filename")
    index_version = body.get("index_version")
    sheet_name = body.get("sheet")
    file_bytes_b64 = body.get("file_bytes")
    headers = body.get("headers")
    rows = body.get("rows")
    mutations = body.get("mutations")

    if not all([doc_id, index_version, file_bytes_b64, isinstance(headers, list), isinstance(rows, list), isinstance(mutations, list)]):
        return jsonify({"success": False, "error": "Missing required fields"}), 400

    try:
        # 1. Parse original document bytes
        file_data = base64.b64decode(file_bytes_b64)
        is_csv = filename.lower().endswith(".csv")
        if is_csv or sheet_name == "CSV":
            sheet_name = None

        # 2. Modify specific cells in the parsed object
        if is_csv:
            csv_in = io.StringIO(file_data.decode("utf-8"))
            csv_rows = list(csv.reader(csv_in))
            
            for m in mutations:
                r_idx = int(m["row"]) + 1
                c_idx = int(m["column"])
                if 0 <= r_idx < len(csv_rows) and 0 <= c_idx < len(csv_rows[r_idx]):
                    csv_rows[r_idx][c_idx] = str(m["value"])
            
            csv_out = io.StringIO()
            writer = csv.writer(csv_out)
            writer.writerows(csv_rows)
            new_file_bytes = csv_out.getvalue().encode("utf-8")
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=False)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            for m in mutations:
                r_idx = int(m["row"]) + 2
                c_idx = int(m["column"]) + 1
                ws.cell(row=r_idx, column=c_idx, value=m["value"])

            xlsx_out = io.BytesIO()
            wb.save(xlsx_out)
            new_file_bytes = xlsx_out.getvalue()

        # 3. Retrieve all Chroma chunks for the sheet (both table and paragraph formats)
        where_filter = {
            "$and": [
                {"doc_id": doc_id},
                {"index_version": index_version}
            ]
        }
        if sheet_name:
            where_filter["$and"].append({"sheet": sheet_name})

        chroma_res = collection.get(where=where_filter, include=["metadatas", "documents"])
        metadatas = chroma_res.get("metadatas") or []
        documents = chroma_res.get("documents") or []
        ids = chroma_res.get("ids") or []

        table_chunks = []
        paragraph_chunks = []

        for cid, meta, doc_text in zip(ids, metadatas, documents):
            if meta.get("is_table") or meta.get("chunk_type") == "table":
                table_chunks.append({"id": cid, "meta": meta, "text": doc_text})
            elif meta.get("chunk_type") == "paragraph":
                paragraph_chunks.append({"id": cid, "meta": meta, "text": doc_text})

        # 4. Map mutations to unique affected table chunk indices
        affected_indices = set()
        chunk_map = {}
        for tc in table_chunks:
            cid = tc["id"]
            meta = tc["meta"]
            doc_text = tc["text"]
            row_start = meta.get("row_start")
            row_end = meta.get("row_end")
            chunk_num = meta.get("chunk")
            if row_start is not None and row_end is not None:
                chunk_map[chunk_num] = {
                    "id": cid,
                    "row_start": row_start,
                    "row_end": row_end,
                    "meta": meta,
                    "doc_text": doc_text
                }
                for m in mutations:
                    row_idx = int(m["row"])
                    if row_start <= row_idx < row_end:
                        affected_indices.add(chunk_num)

        # 5. Regenerate, compute checksums, and update modified table chunks
        updated_chunks = []
        for chunk_num in affected_indices:
            info = chunk_map[chunk_num]
            cid = info["id"]
            row_start = info["row_start"]
            row_end = info["row_end"]
            meta = info["meta"]

            subset = rows[row_start:row_end]
            new_flat_text = flatten_table_for_embedding(sheet=sheet_name, headers=headers, rows=subset)
            new_hash = compute_chunk_hash(new_flat_text)

            old_hash = meta.get("chunk_hash")
            if old_hash == new_hash:
                logger.info("[SyncTable] Chunk %s hash matches, skipping re-embedding", cid)
                continue

            logger.info("[SyncTable] Chunk %s modified, re-embedding", cid)

            embed_header = _build_contextual_header(filename, sheet_name=sheet_name, include_filename=False)
            emb = embed_document(text=new_flat_text, title=filename, context=embed_header)
            if not emb:
                raise Exception(f"Failed to generate embedding for chunk {chunk_num}")

            meta["chunk_hash"] = new_hash
            collection.upsert(ids=[cid], embeddings=[emb], documents=[new_flat_text], metadatas=[meta])

            with _bm25_lock:
                entry = _bm25_cache.get((doc_id, index_version))
                if entry and cid in entry["chunk_ids"]:
                    idx = entry["chunk_ids"].index(cid)
                    entry["texts"][idx] = new_flat_text

            updated_chunks.append({
                "chunk_id": cid,
                "chunk_index": chunk_num,
                "text": new_flat_text
            })

        # 6. Rebuild and update paragraph chunks for this sheet
        paragraph_deletions = []
        paragraph_insertions = []

        if paragraph_chunks:
            from utils.table_extraction import _raw_text_from_table, tables_to_scan_text
            from indexing.chunking import chunk_text

            updated_table = {
                "sheet": sheet_name,
                "headers": headers,
                "rows": rows,
                "raw_text": _raw_text_from_table(headers, rows, sheet_name),
                "flattened_text": flatten_table_for_embedding(sheet=sheet_name, headers=headers, rows=rows)
            }
            body = tables_to_scan_text([updated_table])
            new_paragraphs = chunk_text(body)

            # Align paragraph chunks in index order
            paragraph_chunks.sort(key=lambda x: x["meta"].get("chunk", 0))
            old_indices = sorted([pc["meta"].get("chunk") for pc in paragraph_chunks])

            # Build chunk index map for O(1) lookups instead of O(n) linear scans
            paragraph_by_chunk = {pc["meta"].get("chunk"): pc for pc in paragraph_chunks}

            # Resolve maximum chunk index in the entire document version to allocate collision-free indices
            all_meta_res = collection.get(
                where={"$and": [{"doc_id": doc_id}, {"index_version": index_version}]},
                include=["metadatas"]
            )
            all_metas = all_meta_res.get("metadatas") or []
            max_chunk_index = max([m.get("chunk") for m in all_metas] or [-1])

            for i, new_txt in enumerate(new_paragraphs):
                if i < len(old_indices):
                    # Reuse existing chunk indices and update in-place
                    chunk_num = old_indices[i]
                    pc = paragraph_by_chunk[chunk_num]
                    cid = pc["id"]
                    meta = pc["meta"]
                    new_hash = compute_chunk_hash(new_txt)

                    old_hash = meta.get("chunk_hash")
                    if old_hash == new_hash:
                        continue

                    logger.info("[SyncTable] Paragraph chunk %s modified, re-embedding", cid)
                    embed_header = _build_contextual_header(
                        filename=filename,
                        section=meta.get("section"),
                        subsection=meta.get("subsection"),
                        start_page=meta.get("start_page"),
                        end_page=meta.get("end_page"),
                        sheet_name=sheet_name,
                        include_filename=False
                    )
                    emb = embed_document(text=new_txt, title=filename, context=embed_header)
                    if not emb:
                        raise Exception(f"Failed to generate embedding for paragraph chunk {cid}")

                    meta["chunk_hash"] = new_hash
                    collection.upsert(ids=[cid], embeddings=[emb], documents=[new_txt], metadatas=[meta])

                    with _bm25_lock:
                        entry = _bm25_cache.get((doc_id, index_version))
                        if entry and cid in entry["chunk_ids"]:
                            idx = entry["chunk_ids"].index(cid)
                            entry["texts"][idx] = new_txt

                    updated_chunks.append({
                        "chunk_id": cid,
                        "chunk_index": chunk_num,
                        "text": new_txt
                    })
                else:
                    # Allocate a brand-new sequential index to prevent unique key constraint collisions in Mongo
                    max_chunk_index += 1
                    chunk_num = max_chunk_index
                    cid = f"{doc_id}:{index_version}:{chunk_num}"
                    new_hash = compute_chunk_hash(new_txt)

                    base_meta = paragraph_chunks[0]["meta"]
                    meta = {
                        "doc_id": doc_id,
                        "index_version": index_version,
                        "chunk": chunk_num,
                        "filename": filename,
                        "source_type": base_meta.get("source_type", "xlsx"),
                        "section": base_meta.get("section", ""),
                        "subsection": base_meta.get("subsection", ""),
                        "heading_level": -1,
                        "start_page": 1,
                        "end_page": 1,
                        "section_index": base_meta.get("section_index", 0),
                        "chunk_index": chunk_num,
                        "embedding_model": base_meta.get("embedding_model", EMBED_MODEL),
                        "pipeline_version": base_meta.get("pipeline_version", INDEX_PIPELINE_VERSION),
                        "chunking_version": base_meta.get("chunking_version", CHUNKING_VERSION),
                        "chunk_type": "paragraph",
                        "paragraph_count": 1,
                        "token_count": estimate_token_count(new_txt),
                        "indexed_at": base_meta.get("indexed_at"),
                        "chunk_header": base_meta.get("chunk_header", ""),
                        "chunk_hash": new_hash
                    }
                    if sheet_name:
                        meta["sheet"] = sheet_name
                    if base_meta.get("file_hash"):
                        meta["file_hash"] = base_meta.get("file_hash")

                    embed_header = _build_contextual_header(filename, sheet_name=sheet_name, include_filename=False)
                    emb = embed_document(text=new_txt, title=filename, context=embed_header)
                    if not emb:
                        raise Exception(f"Failed to generate embedding for new paragraph chunk {cid}")

                    collection.add(ids=[cid], embeddings=[emb], documents=[new_txt], metadatas=[meta])

                    with _bm25_lock:
                        entry = _bm25_cache.get((doc_id, index_version))
                        if entry:
                            entry["chunk_ids"].append(cid)
                            entry["texts"].append(new_txt)

                    paragraph_insertions.append({
                        "chunk_index": chunk_num,
                        "text": new_txt
                    })

            # Handle deletions of excess paragraph chunks if count decreased
            if len(new_paragraphs) < len(old_indices):
                deleted_indices = old_indices[len(new_paragraphs):]
                for chunk_num in deleted_indices:
                    pc = paragraph_by_chunk[chunk_num]
                    cid = pc["id"]
                    
                    collection.delete(ids=[cid])

                    with _bm25_lock:
                        entry = _bm25_cache.get((doc_id, index_version))
                        if entry and cid in entry["chunk_ids"]:
                            idx = entry["chunk_ids"].index(cid)
                            entry["chunk_ids"].pop(idx)
                            entry["texts"].pop(idx)

                    paragraph_deletions.append(chunk_num)

        if updated_chunks or paragraph_insertions or paragraph_deletions:
            debounce_bm25_rebuild(doc_id, index_version)

        new_file_b64 = base64.b64encode(new_file_bytes).decode("utf-8")

        return jsonify({
            "success": True,
            "file_bytes": new_file_b64,
            "updated_chunks": updated_chunks,
            "recreated_paragraphs": {
                "delete_chunk_indices": paragraph_deletions,
                "new_chunks": paragraph_insertions
            }
        })

    except Exception as err:
        logger.error("[SyncTable] Failed cell edit synchronization: %s", err)
        return jsonify({"success": False, "error": str(err)}), 500
