import os
import base64
import csv
import io
from unittest.mock import patch, MagicMock

os.environ["SERVICE_TOKEN"] = "test_service_token"

import pytest
from main import app
from openpyxl import Workbook


@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client


@patch("features.table_edit.collection")
@patch("features.table_edit.embed_document")
def test_sync_table_csv(mock_embed, mock_collection, client):
    # Mock Chroma collection get
    mock_collection.get.return_value = {
        "ids": ["doc_id:v1:0"],
        "metadatas": [{"row_start": 0, "row_end": 5, "chunk": 0, "is_table": True, "chunk_hash": "old_hash"}],
        "documents": ["Header1Header2\nRow1Val1Row1Val2"]
    }
    mock_embed.return_value = [0.1] * 768

    # Create dummy CSV
    csv_out = io.StringIO()
    writer = csv.writer(csv_out)
    writer.writerow(["Name", "Age"])
    writer.writerow(["Alice", "25"])
    file_bytes = csv_out.getvalue().encode("utf-8")
    file_b64 = base64.b64encode(file_bytes).decode("utf-8")

    payload = {
        "doc_id": "doc123",
        "filename": "file.csv",
        "index_version": "v1",
        "sheet": None,
        "file_bytes": file_b64,
        "headers": ["Name", "Age"],
        "rows": [["Alice", "25"]],
        "mutations": [
            {"row": 0, "column": 1, "value": "26"}
        ]
    }

    resp = client.post(
        "/api/internal/document/sync-table",
        json=payload,
        headers={"x-service-token": "test_service_token"}
    )

    assert resp.status_code == 200
    res_data = resp.get_json()
    assert res_data["success"] is True
    assert "file_bytes" in res_data
    assert len(res_data["updated_chunks"]) == 1
    assert res_data["updated_chunks"][0]["chunk_id"] == "doc_id:v1:0"
    assert res_data["updated_chunks"][0]["chunk_index"] == 0

    # Decode returned file and verify cell was updated
    returned_bytes = base64.b64decode(res_data["file_bytes"])
    csv_in = io.StringIO(returned_bytes.decode("utf-8"))
    csv_rows = list(csv.reader(csv_in))
    assert csv_rows[1][1] == "26"


@patch("features.table_edit.collection")
@patch("features.table_edit.embed_document")
def test_sync_table_xlsx(mock_embed, mock_collection, client):
    # Mock Chroma collection get
    mock_collection.get.return_value = {
        "ids": ["doc_id:v1:0"],
        "metadatas": [{"row_start": 0, "row_end": 5, "chunk": 0, "is_table": True, "chunk_hash": "old_hash"}],
        "documents": ["Header1Header2\nRow1Val1Row1Val2"]
    }
    mock_embed.return_value = [0.1] * 768

    # Create dummy Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Name")
    ws.cell(row=1, column=2, value="Age")
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=2, value="25")

    xlsx_out = io.BytesIO()
    wb.save(xlsx_out)
    file_bytes = xlsx_out.getvalue()
    file_b64 = base64.b64encode(file_bytes).decode("utf-8")

    payload = {
        "doc_id": "doc123",
        "filename": "file.xlsx",
        "index_version": "v1",
        "sheet": "Sheet",
        "file_bytes": file_b64,
        "headers": ["Name", "Age"],
        "rows": [["Alice", "25"]],
        "mutations": [
            {"row": 0, "column": 1, "value": "26"}
        ]
    }

    resp = client.post(
        "/api/internal/document/sync-table",
        json=payload,
        headers={"x-service-token": "test_service_token"}
    )

    assert resp.status_code == 200
    res_data = resp.get_json()
    assert res_data["success"] is True
    assert "file_bytes" in res_data
    assert len(res_data["updated_chunks"]) == 1


@patch("features.table_edit.collection")
@patch("features.table_edit.embed_document")
def test_sync_table_checksum_no_change(mock_embed, mock_collection, client):
    # Mock Chroma collection get
    # Note: we set chunk_hash to match the new computed hash (which we will compute from "Name = Alice\nAge = 25")
    from indexing.pipeline import compute_chunk_hash
    text = "Row:\nName = Alice\nAge = 25"
    expected_hash = compute_chunk_hash(text)

    mock_collection.get.return_value = {
        "ids": ["doc_id:v1:0"],
        "metadatas": [{"row_start": 0, "row_end": 5, "chunk": 0, "is_table": True, "chunk_hash": expected_hash}],
        "documents": [text]
    }

    # Create dummy CSV
    csv_out = io.StringIO()
    writer = csv.writer(csv_out)
    writer.writerow(["Name", "Age"])
    writer.writerow(["Alice", "25"])
    file_bytes = csv_out.getvalue().encode("utf-8")
    file_b64 = base64.b64encode(file_bytes).decode("utf-8")

    payload = {
        "doc_id": "doc123",
        "filename": "file.csv",
        "index_version": "v1",
        "sheet": None,
        "file_bytes": file_b64,
        "headers": ["Name", "Age"],
        "rows": [["Alice", "25"]],
        "mutations": [
            {"row": 0, "column": 1, "value": "25"}  # no change mutation
        ]
    }

    resp = client.post(
        "/api/internal/document/sync-table",
        json=payload,
        headers={"x-service-token": "test_service_token"}
    )

    assert resp.status_code == 200
    res_data = resp.get_json()
    assert res_data["success"] is True
    # Verify no update happened because hash matches
    assert len(res_data["updated_chunks"]) == 0
    mock_embed.assert_not_called()
    mock_collection.upsert.assert_not_called()
